"""
Журнал успеваемости студентов
"""

import json
import os
from dataclasses import dataclass, asdict
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

DB_FILE = "gradebook_data.json"
EXCEL_FILE = "gradebook.xlsx"

MIN_ZAD_LIMIT = 2
APP_VERSION = "Альфа 1.0"


@dataclass
class Student:
    n: str
    f: str
    stud_ball: List[int]

    @property
    def average(self) -> float:
        return sum(self.stud_ball) / len(self.stud_ball) if self.stud_ball else 0.0


class GradeBook:
    def __init__(self):
        self.kol_zad = 0
        self.spisok_stud: List[Student] = []
        self.load_from_json()

    # ---------- data ----------
    def set_config(self, kol_zad: int):
        if kol_zad < MIN_ZAD_LIMIT:
            raise ValueError("Минимум 2 задания")
        if self.kol_zad == 0:
            self.kol_zad = kol_zad

    def add_student(self, st: Student):
        self._validate_student(st)
        self.spisok_stud.append(st)
        self.save_to_json()

    def update_student(self, index: int, st: Student):
        self._validate_student(st)
        self.spisok_stud[index] = st
        self.save_to_json()

    def _validate_student(self, st: Student):
        if len(st.stud_ball) != self.kol_zad:
            raise ValueError("Количество оценок не совпадает с количеством заданий")
        if any(g < 1 or g > 5 for g in st.stud_ball):
            raise ValueError("Оценки должны быть от 1 до 5")

    def reset_data(self):
        self.kol_zad = 0
        self.spisok_stud.clear()
        if os.path.exists(DB_FILE):
            os.remove(DB_FILE)

    def save_to_json(self):
        with open(DB_FILE, "w", encoding="utf-8") as f:
            json.dump(
                {"kol_zad": self.kol_zad, "students": [asdict(s) for s in self.spisok_stud]},
                f, ensure_ascii=False, indent=4
            )

    def load_from_json(self):
        if not os.path.exists(DB_FILE):
            return
        try:
            with open(DB_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            self.kol_zad = data.get("kol_zad", 0)
            self.spisok_stud = [Student(**s) for s in data.get("students", [])]
        except Exception:
            pass

    # ---------- excel ----------
    def export_excel(self, filename: Optional[str] = None):
        if not self.spisok_stud:
            raise ValueError("Список пуст")

        filename = filename or EXCEL_FILE
        wb = Workbook()
        ws = wb.active
        ws.title = "Успеваемость"

        headers = (
            ["№", "Имя", "Фамилия"]
            + [f"Задание {i+1}" for i in range(self.kol_zad)]
            + ["Средний балл"]
        )
        ws.append(headers)

        for i, s in enumerate(self.spisok_stud, start=1):
            ws.append([i, s.n, s.f, *s.stud_ball, round(s.average, 2)])

        # стиль
        header_fill = PatternFill("solid", fgColor="1F4E79")
        for c in ws[1]:
            c.fill = header_fill
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
            c.border = Border(*(Side(style="thin"),) * 4)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        avg_col = ws.max_column
        avg_range = f"{get_column_letter(avg_col)}2:{get_column_letter(avg_col)}{ws.max_row}"

        ws.conditional_formatting.add(
            avg_range, CellIsRule(operator="lessThan", formula=["3"], fill=PatternFill("solid", fgColor="FFC7CE"))
        )
        ws.conditional_formatting.add(
            avg_range, CellIsRule(operator="greaterThanOrEqual", formula=["4.5"], fill=PatternFill("solid", fgColor="C6EFCE"))
        )

        ws.append([])
        ws.append([f"Экспортировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"])

        wb.save(filename)
